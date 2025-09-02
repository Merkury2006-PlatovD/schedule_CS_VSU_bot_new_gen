import re
import openpyxl
from openpyxl.cell.cell import MergedCell

from src.parser_service.util.error import ScheduleParserFindError, NotFoundListError


class ScheduleParser:
    time_lessons = ['8:00-9:35', '9:45-11:20', '11:30-13:05', '13:25-15:00',
                    '15:10-16:45', '16:55-18:30', '18:40-20:00', '20:10-21:30']
    filename: str

    def __init__(self, filename):
        self.filename = filename
        self.all_courses = None
        self.sheet = None
        self.wb = None

    def refresh_workbook(self):
        self.wb = openpyxl.load_workbook(self.filename)
        self.sheet = self.__find_required_list()
        print(f"Бот использует лист под названием: {self.sheet.title}")
        self.all_courses = self._parse_headers()

    def __find_required_list(self):
        for sheet in self.wb:
            print(sheet.title)
            if re.search(r'бак', sheet.title, re.IGNORECASE):
                return sheet
        raise NotFoundListError()

    def _get_merged_cell_value(self, cell):
        if isinstance(cell, MergedCell):
            for merged_range in self.sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return self.sheet[merged_range.start_cell.coordinate].value
        return cell.value

    def _parse_headers(self):
        headers = []
        for row in self.sheet.iter_rows(max_row=2, min_col=3):
            values = []
            for cell in row:
                val = self._get_merged_cell_value(cell)

                if val:
                    matcher = re.search(r"\d+", val)
                    if matcher is not None:
                        values.append(matcher.group())
                    else:
                        values.append("NN")
            headers.append(values)
        return headers

    def find_required_col(self, course, group, subgroup):
        for i in range(len(self.all_courses[0])):
            if self.all_courses[0][i] == str(course) and self.all_courses[1][i] == str(group):
                if subgroup == 1:
                    return i + 3  # плюсую значение чтобы дойти до нужного столбца тк первые два столбца с днями недели и временем ну и индексация с нуля
                subgroup -= 1
        raise ScheduleParserFindError

    def get_lessons_on_day(self, column, day, week):
        time_lesson_gen = (x for x in self.time_lessons)
        col_values = [self._get_merged_cell_value(row[column - 1]) for row in self.sheet.iter_rows(min_row=5)]
        skip = day * 16
        out_schedule = {}

        for i in range(skip + week + day, len(col_values), 2):
            try:
                out_schedule[next(time_lesson_gen)] = col_values[i]
            except StopIteration:
                break

        return out_schedule

# parser = ScheduleParser('schedule.xlsx')
# column = parser.find_required_col(1, 2, 1)
# parser.get_lessons_on_day(column, 0, 0)
# неделя 0 - числитель 1 - знаменатель
# day 0 - понедельник 1 - вторник и тд
# курс группа подгруппа как есть
